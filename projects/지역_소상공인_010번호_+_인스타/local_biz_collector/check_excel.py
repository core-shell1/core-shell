import openpyxl, os, glob

export_dir = "Z:/data/exports"
files = sorted(glob.glob(f"{export_dir}/*.xlsx"))

for f in files:
    try:
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        total = ws.max_row - 1

        phone_col = None
        insta_col = None
        for i, h in enumerate(headers):
            if h and ('010' in str(h) or '전화' in str(h)):
                phone_col = i + 1
            if h and '인스타' in str(h):
                insta_col = i + 1

        c010 = 0
        cinsta = 0
        if phone_col:
            c010 = sum(1 for r in range(2, ws.max_row+1)
                       if str(ws.cell(r, phone_col).value or '').startswith('010'))
        if insta_col:
            cinsta = sum(1 for r in range(2, ws.max_row+1)
                         if ws.cell(r, insta_col).value)

        print(f"{os.path.basename(f)}: 총 {total}건 | 010={c010}건 | 인스타={cinsta}건")
        print(f"  컬럼: {headers[:8]}")
        # 샘플 2행
        for r in range(2, min(4, ws.max_row+1)):
            row = [ws.cell(r, c).value for c in range(1, min(7, ws.max_column+1))]
            print(f"  샘플: {row}")
        print()
    except Exception as e:
        print(f"{os.path.basename(f)}: 오류 {e}")
