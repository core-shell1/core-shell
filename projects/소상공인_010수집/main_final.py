"""
소상공인 010번호 수집기 — 최종본
=================================
소스 (전부 동시 실행):
  1. 카카오맵 API   — JSONP, asyncio 10개 동시, Playwright 없음 (가장 많이 나옴)
  2. 네이버맵 API   — map.naver.com allSearch 인터셉트, 구조화 데이터
  3. 당근마켓       — /local-profile/ 3개 브라우저 병렬
  4. 네이버 웹검색  — 4개 브라우저 병렬
  5. DuckDuckGo     — 2개 브라우저 병렬

제거: 네이버블로그 (6개밖에 안 나옴), 다음검색 (네이버와 중복)

사용법:
  python main_final.py            ← 양주
  python main_final.py 의정부
  python main_final.py 포천 --verify
"""

import asyncio
import json
import re
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ─────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────
REGION    = next((a for a in sys.argv[1:] if not a.startswith("--")), "양주")
DO_VERIFY = "--verify" in sys.argv
ONLY      = next((a.split("=")[1] for a in sys.argv if a.startswith("--only=")), None)

# 지역별 서브지역 (카카오맵 검색에 전부 사용)
SUB_AREAS = {
    "양주": [
        "양주", "양주 회천", "양주 백석", "양주 덕정", "양주 덕계",
        "양주 옥정", "양주 장흥", "양주 광적", "양주 남면", "양주 은현",
    ],
    "의정부": [
        "의정부", "의정부 가능", "의정부 녹양", "의정부 신곡",
        "의정부 호원", "의정부 장암", "의정부 민락", "의정부 흥선",
    ],
    "포천": [
        "포천", "포천 소흘", "포천 군내", "포천 동면",
        "포천 신북", "포천 가산", "포천 일동", "포천 영중",
    ],
}

DESKTOP_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)
MOBILE_UA = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
)

# ── 카카오맵 카테고리 (39개) ──────────────────
KAKAO_CATS = [
    "치킨", "피자", "족발", "삼겹살", "분식", "중국집", "일식", "돈까스",
    "국밥", "냉면", "빵집", "도시락", "맛집", "식당", "떡볶이", "순대국",
    "미용실", "네일", "헬스장", "필라테스", "피부관리", "왁싱", "속눈썹",
    "카페", "커피숍", "버블티",
    "학원", "공부방", "과외", "레슨",
    "세탁소", "청소업체", "인테리어", "수리", "이사", "꽃집",
    "반려동물", "동물병원", "애견미용",
]

# ── 네이버맵 / 당근 / 웹 검색 카테고리 ────────
WEB_CATS = [
    "음식점", "카페", "미용실", "네일", "헬스장", "필라테스",
    "학원", "병원", "치과", "한의원", "피부과",
    "세탁소", "꽃집", "베이커리", "빵집",
    "치킨", "피자", "족발", "보쌈", "분식",
    "인테리어", "반려동물", "애견미용",
    "부동산", "사진관", "게스트하우스",
]

NAVER_MAP_CATS = [
    "카페", "음식점", "미용실", "네일샵", "헬스장", "필라테스",
    "학원", "병원", "치과", "세탁소", "꽃집",
    "치킨", "피자", "족발", "분식",
    "인테리어", "반려동물",
]

FRANCHISE_EXCLUDE = {
    "스타벅스", "이마트", "홈플러스", "롯데마트", "코스트코", "다이소",
    "맥도날드", "버거킹", "롯데리아", "KFC", "서브웨이", "맘스터치",
    "GS25", "CU", "세븐일레븐", "미니스톱", "이마트24",
    "올리브영", "유니클로", "자라", "H&M", "크린토피아",
    "bbq", "굽네치킨", "교촌치킨", "bhc",
}

PHONE_RE    = re.compile(r'01[016789][-.\s]?\d{3,4}[-.\s]?\d{4}')
PLACE_ID_RE = re.compile(r'/place/(\d{6,})')

_lock    = asyncio.Lock()
_records: dict[str, dict] = {}

# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def normalize(raw: str) -> str:
    d = re.sub(r'[^\d]', '', raw)
    if len(d) == 11: return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) == 10: return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    return raw

def is_010(raw: str) -> bool:
    d = re.sub(r'[^\d]', '', raw)
    return d.startswith('010') and len(d) == 11

def is_franchise(name: str) -> bool:
    return any(f in name for f in FRANCHISE_EXCLUDE)

async def add(phone_raw: str, name: str, source: str, cat: str = ""):
    if not is_010(phone_raw): return
    phone = normalize(phone_raw)
    if is_franchise(name): return
    async with _lock:
        if phone not in _records:
            _records[phone] = {
                "phone": phone, "name": name, "category": cat,
                "source": source, "verified": False,
                "biz_name": "", "address": "", "naver_url": "",
            }


# ─────────────────────────────────────────────
# 1. 카카오맵 API (HTTP 10개 병렬)
# ─────────────────────────────────────────────
async def _kakao_one(client: httpx.AsyncClient, sem: asyncio.Semaphore,
                     query: str, page: int, cat: str) -> int:
    url = (
        "https://search.map.kakao.com/mapsearch/map.daum"
        f"?callback=cb&q={urllib.parse.quote(query)}&page={page}&size=15&sort=0"
    )
    hdrs = {"User-Agent": DESKTOP_UA, "Referer": "https://map.kakao.com/",
            "Accept-Language": "ko-KR,ko;q=0.9"}
    async with sem:
        try:
            resp = await client.get(url, headers=hdrs, timeout=10)
            js = re.sub(r'^/\*\*/cb\(', '', resp.text.strip())
            js = re.sub(r'\);\s*$', '', js)
            places = json.loads(js).get("place", [])
        except Exception:
            return 0
        await asyncio.sleep(0.2)

    cnt = 0
    for p in places:
        tel = p.get("tel", "")
        if is_010(tel):
            before = len(_records)
            await add(tel, p.get("name", ""), "카카오맵", cat)
            if len(_records) > before: cnt += 1
    return cnt


async def scrape_kakao(region: str) -> int:
    areas = SUB_AREAS.get(region, [region])
    sem = asyncio.Semaphore(10)
    tasks = []
    async with httpx.AsyncClient() as client:
        for area in areas:
            for cat in KAKAO_CATS:
                q = f"{area} {cat}"
                for pg in range(1, 16):  # 15페이지 × 15개 = 카테고리당 최대 225개
                    tasks.append(_kakao_one(client, sem, q, pg, cat))
        results = await asyncio.gather(*tasks)
    return sum(results)


# ─────────────────────────────────────────────
# 2. 네이버맵 API 인터셉트 (allSearch 응답 가로채기)
# ─────────────────────────────────────────────
async def _nmap_search(page, query: str) -> int:
    pending_responses = []

    # 동기 핸들러로 response 객체만 저장 (await 없음)
    def on_response(resp):
        if "allSearch" in resp.url:
            pending_responses.append(resp)

    page.on("response", on_response)
    try:
        await page.goto(
            f"https://map.naver.com/p/search/{urllib.parse.quote(query)}",
            timeout=25000
        )
        await page.wait_for_load_state("networkidle", timeout=20000)
        await asyncio.sleep(2.0)
        # 스크롤로 추가 로드
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await asyncio.sleep(1.5)
    except Exception:
        pass
    finally:
        page.remove_listener("response", on_response)

    cnt = 0
    for resp in pending_responses:
        try:
            data = await resp.json()
            items = (data.get("result") or {}).get("place", {})
            if isinstance(items, dict):
                items = items.get("list", [])
            elif not isinstance(items, list):
                items = []
            for item in items:
                tel = item.get("tel", "")
                if is_010(tel):
                    before = len(_records)
                    await add(tel, item.get("name", ""), "네이버맵", "")
                    if len(_records) > before: cnt += 1
        except Exception:
            pass
    return cnt


async def scrape_naver_map(short: str) -> int:
    total = 0
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx  = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR",
                                          viewport={"width": 1280, "height": 800})
        ctx2 = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR",
                                          viewport={"width": 1280, "height": 800})
        page1 = await ctx.new_page()
        page2 = await ctx2.new_page()

        mid = len(NAVER_MAP_CATS) // 2
        cats_a = NAVER_MAP_CATS[:mid]
        cats_b = NAVER_MAP_CATS[mid:]

        async def worker(page, cats):
            n = 0
            for cat in cats:
                n += await _nmap_search(page, f"{short} {cat}")
                await asyncio.sleep(0.5)
            return n

        r1, r2 = await asyncio.gather(worker(page1, cats_a), worker(page2, cats_b))
        total = r1 + r2
        await browser.close()
    return total


# ─────────────────────────────────────────────
# 3. 당근마켓 (3개 브라우저 병렬)
# ─────────────────────────────────────────────
async def _daangn_worker(page, cats: list, short: str) -> int:
    cnt = 0
    for cat in cats:
        q = f"{short} {cat}"
        url = f"https://www.daangn.com/kr/local-profile/s/?search={urllib.parse.quote(q)}"
        try:
            await page.goto(url, timeout=25000)
            await page.wait_for_load_state("networkidle", timeout=20000)
            await asyncio.sleep(1.5)

            links = await page.evaluate("""() =>
                Array.from(document.querySelectorAll('a[href*="/local-profile/"]'))
                    .map(a => a.href)
                    .filter(h => !/\\/local-profile\\/s\\//.test(h))
                    .filter((v,i,a) => a.indexOf(v)===i)
                    .slice(0,8)
            """)

            for link in links:
                try:
                    await page.goto(link, timeout=18000)
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    await asyncio.sleep(0.8)
                    data = await page.evaluate("""() => {
                        const h = document.querySelector('h1,h2')?.textContent?.trim()||''
                        const body = document.body.innerText
                        const phones = body.match(/010[-\\s]?\\d{3,4}[-\\s]?\\d{4}/g)||[]
                        const tels = Array.from(document.querySelectorAll('a[href^="tel:"]'))
                            .map(a=>a.href.replace('tel:','').trim())
                        return {name:h, phones, tels}
                    }""")
                    for ph in list(set(data["phones"] + [t for t in data["tels"] if is_010(t)])):
                        before = len(_records)
                        await add(ph, data["name"], "당근마켓", cat)
                        if len(_records) > before: cnt += 1
                except Exception:
                    pass
                await asyncio.sleep(0.5)
        except Exception:
            pass
        await asyncio.sleep(0.8)
    return cnt


async def scrape_daangn(short: str) -> int:
    # 단일 브라우저 순차 실행 — 병렬보다 안정적, 원본 방식
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx  = await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
        page = await ctx.new_page()
        r = await _daangn_worker(page, WEB_CATS, short)
        await browser.close()
    return r


# ─────────────────────────────────────────────
# 4. 네이버 웹검색 (4개 브라우저 병렬)
# ─────────────────────────────────────────────
async def _naver_web_worker(page, cats: list, short: str) -> int:
    cnt = 0
    for cat in cats:
        for suffix in ["010", "연락처"]:
            q = f"{short} {cat} {suffix}"
            for start in [1, 11, 21, 31, 41]:  # 5페이지
                url = f"https://search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=web&start={start}"
                try:
                    await page.goto(url, timeout=18000)
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    await asyncio.sleep(0.3)
                    text = await page.inner_text("body")
                    found = 0
                    for m in PHONE_RE.finditer(text):
                        if is_010(m.group(0)):
                            before = len(_records)
                            await add(m.group(0), cat, "네이버웹", cat)
                            if len(_records) > before:
                                cnt += 1
                                found += 1
                    # 새 번호 없으면 다음 페이지 안 봄
                    if found == 0 and start > 1:
                        break
                except Exception:
                    pass
                await asyncio.sleep(0.2)
    return cnt


async def scrape_naver_web(short: str) -> int:
    n = len(WEB_CATS)
    parts = [WEB_CATS[i*n//4:(i+1)*n//4] for i in range(4)]

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctxs  = [await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
                 for _ in range(4)]
        pages = [await ctx.new_page() for ctx in ctxs]
        r = await asyncio.gather(*[_naver_web_worker(pages[i], parts[i], short)
                                    for i in range(4)])
        await browser.close()
    return sum(r)


# ─────────────────────────────────────────────
# 5. DuckDuckGo (ddgs 라이브러리 + run_in_executor)
# ─────────────────────────────────────────────
def _ddg_sync(queries: list[str]) -> list[tuple[str, str]]:
    """동기 DDG 검색 — run_in_executor로 호출"""
    from ddgs import DDGS
    results = []
    ddgs = DDGS()
    for q in queries:
        try:
            for r in ddgs.text(q, region="kr-kr", max_results=20):
                text = r.get("title","") + " " + r.get("body","")
                results.append((q, text))
        except Exception:
            pass
    return results


async def scrape_ddg(short: str) -> int:
    queries = [f"{short} {cat} 010" for cat in WEB_CATS]
    loop = asyncio.get_event_loop()
    # 절반씩 두 스레드에서 병렬 실행
    half = len(queries) // 2
    res_a, res_b = await asyncio.gather(
        loop.run_in_executor(None, _ddg_sync, queries[:half]),
        loop.run_in_executor(None, _ddg_sync, queries[half:]),
    )
    cnt = 0
    for q, text in res_a + res_b:
        cat = q.split(" ")[-2] if len(q.split()) >= 2 else ""
        for m in PHONE_RE.finditer(text):
            if is_010(m.group(0)):
                before = len(_records)
                await add(m.group(0), cat, "DuckDuckGo", cat)
                if len(_records) > before: cnt += 1
    return cnt


# ─────────────────────────────────────────────
# 6. 네이버 place 검색 (where=place, 업체 카드에서 직접 추출)
# ─────────────────────────────────────────────
async def _naver_place_worker(page, cats: list, short: str) -> int:
    cnt = 0
    for cat in cats:
        q = f"{short} {cat}"
        url = f"https://search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=place"
        try:
            await page.goto(url, timeout=18000)
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await asyncio.sleep(0.5)
            # 스크롤로 더 로드
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await asyncio.sleep(0.4)
            text = await page.inner_text("body")
            # 업체명도 같이 추출 시도
            items = await page.evaluate("""() => {
                const results = []
                // 네이버 place 결과 카드
                document.querySelectorAll('li[class*="place_item"], li[class*="PlaceItem"], div[class*="place_item"]').forEach(el => {
                    const name = el.querySelector('[class*="name"], [class*="title"], h3, h2')?.textContent?.trim() || ''
                    const phones = (el.innerText.match(/010[-\\s]?\\d{3,4}[-\\s]?\\d{4}/g) || [])
                    const tels = Array.from(el.querySelectorAll('a[href^="tel:"]')).map(a => a.href.replace('tel:',''))
                    phones.concat(tels).forEach(p => results.push({name, phone: p}))
                })
                return results
            }""")
            # items에서 추출
            for item in items:
                ph = item.get("phone", "")
                nm = item.get("name", "")
                if is_010(ph):
                    before = len(_records)
                    await add(ph, nm, "네이버place", cat)
                    if len(_records) > before: cnt += 1
            # fallback: body 전체 텍스트에서 정규식
            for m in PHONE_RE.finditer(text):
                if is_010(m.group(0)):
                    before = len(_records)
                    await add(m.group(0), cat, "네이버place", cat)
                    if len(_records) > before: cnt += 1
        except Exception:
            pass
        await asyncio.sleep(0.3)
    return cnt


async def scrape_naver_place(short: str) -> int:
    n = len(WEB_CATS)
    parts = [WEB_CATS[i*n//4:(i+1)*n//4] for i in range(4)]

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctxs  = [await browser.new_context(user_agent=DESKTOP_UA, locale="ko-KR")
                 for _ in range(4)]
        pages = [await ctx.new_page() for ctx in ctxs]
        r = await asyncio.gather(*[_naver_place_worker(pages[i], parts[i], short)
                                    for i in range(4)])
        await browser.close()
    return sum(r)


# ─────────────────────────────────────────────
# 네이버지도 검증 (옵션)
# ─────────────────────────────────────────────
async def verify_all(records: list, region: str):
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx   = await browser.new_context(user_agent=MOBILE_UA, locale="ko-KR")
        page  = await ctx.new_page()
        verified = 0
        for i, rec in enumerate(records, 1):
            q = f"{rec.get('name','')} {region}" if rec.get("name") else rec["phone"]
            try:
                await page.goto(
                    f"https://m.search.naver.com/search.naver?query={urllib.parse.quote(q)}&where=m_local",
                    timeout=18000
                )
                await page.wait_for_load_state("domcontentloaded", timeout=14000)
                await asyncio.sleep(0.5)
                html = await page.content()
                pid  = PLACE_ID_RE.search(html)
                if pid:
                    rec["verified"]  = True
                    rec["naver_url"] = f"https://map.naver.com/p/entry/place/{pid.group(1)}"
                    nm = re.search(r'"placeName"\s*:\s*"([^"]{2,40})"', html)
                    ad = re.search(r'"roadAddress"\s*:\s*"([^"]{5,80})"', html)
                    if nm: rec["biz_name"] = nm.group(1)
                    if ad: rec["address"]  = ad.group(1)
                    verified += 1
            except Exception:
                pass
            if i % 50 == 0:
                print(f"  [{i}/{len(records)}] 확인됨: {verified}개")
            await asyncio.sleep(0.6)
        await browser.close()
    return verified


# ─────────────────────────────────────────────
# 엑셀 저장
# ─────────────────────────────────────────────
def save_excel(records: list, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "010번호"

    hdrs = ["번호", "010번호", "업체명", "업종", "주소", "검증", "네이버플레이스", "출처"]
    hfill = PatternFill("solid", fgColor="1a365d")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left",   vertical="center")

    ws.row_dimensions[1].height = 26
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = ctr

    green  = PatternFill("solid", fgColor="c6efce")
    yellow = PatternFill("solid", fgColor="ffeb9c")

    for idx, rec in enumerate(records, 1):
        row  = idx + 1
        fill = green if rec.get("verified") else yellow
        ws.row_dimensions[row].height = 20
        row_data = [
            idx, rec["phone"],
            rec.get("biz_name") or rec.get("name", ""),
            rec.get("category", ""),
            rec.get("address", ""),
            "✅" if rec.get("verified") else "⚠️",
            rec.get("naver_url", ""),
            rec.get("source", ""),
        ]
        lft_cols = {2,3,4,5,7,8}
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = lft if col in lft_cols else ctr
            if col == 7 and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    for col, w in enumerate([5,16,20,12,35,6,50,12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    Path(path).parent.mkdir(exist_ok=True)
    wb.save(path)


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
async def main():
    global _records
    _records = {}

    region = REGION
    short  = region.replace("경기도","").replace("시","").replace("군","").strip()

    areas = SUB_AREAS.get(region, [region])
    print("=" * 62)
    print(f"  소상공인 010번호 수집기 — 최종본")
    print(f"  지역: {region} ({len(areas)}개 서브지역)  검증: {'ON' if DO_VERIFY else 'OFF'}")
    print(f"  카카오: {len(areas)}지역 × {len(KAKAO_CATS)}카테고리 × 15페이지")
    print(f"  소스: 카카오맵(서브지역) + 당근 + 네이버웹(5p) + 네이버place + DDG")
    print("=" * 62)
    print()

    t0 = datetime.now()

    # ── 소스 선택 실행 ──────────────────────────
    kakao_n = nmap_n = daangn_n = naver_n = ddg_n = place_n = 0

    if ONLY:
        print(f"🚀 단독 실행: {ONLY}")
        if ONLY == "kakao":    kakao_n  = await scrape_kakao(region)
        elif ONLY == "nmap":   nmap_n   = await scrape_naver_map(short)
        elif ONLY == "daangn": daangn_n = await scrape_daangn(short)
        elif ONLY == "web":    naver_n  = await scrape_naver_web(short)
        elif ONLY == "place":  place_n  = await scrape_naver_place(short)
        elif ONLY == "ddg":    ddg_n    = await scrape_ddg(short)
        else: print(f"  알 수 없는 소스: {ONLY}")
    else:
        print("🚀 전체 수집 시작 (모든 소스 동시 실행)...")
        kakao_task   = asyncio.create_task(scrape_kakao(region))
        nmap_task    = asyncio.create_task(scrape_naver_map(short))
        daangn_task  = asyncio.create_task(scrape_daangn(short))
        naver_task   = asyncio.create_task(scrape_naver_web(short))
        ddg_task     = asyncio.create_task(scrape_ddg(short))
        place_task   = asyncio.create_task(scrape_naver_place(short))

        kakao_n, nmap_n, daangn_n, naver_n, ddg_n, place_n = await asyncio.gather(
            kakao_task, nmap_task, daangn_task, naver_task, ddg_task, place_task
        )

    elapsed = (datetime.now() - t0).seconds
    records = list(_records.values())

    print(f"\n{'─'*62}")
    print(f"  수집 완료 ({elapsed//60}분 {elapsed%60}초)")
    print(f"  카카오맵  : +{kakao_n}개")
    print(f"  네이버맵  : +{nmap_n}개")
    print(f"  당근마켓  : +{daangn_n}개")
    print(f"  네이버웹  : +{naver_n}개")
    print(f"  네이버place: +{place_n}개")
    print(f"  DuckDuckGo: +{ddg_n}개")
    print(f"  총 (중복제거): {len(records)}개")
    print(f"{'─'*62}")

    # ── 검증 ────────────────────────────────────
    if DO_VERIFY and records:
        print(f"\n🗺️  네이버지도 검증 중... ({len(records)}개)")
        verified = await verify_all(records, region)
        print(f"  ✅ 확인됨: {verified}개 / 미확인: {len(records)-verified}개")

    # ── 저장 ────────────────────────────────────
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = f"output/{region}_010번호_최종_{ts}.xlsx"
    save_excel(records, out)

    print(f"\n{'='*62}")
    print(f"  완료!  파일: {out}")
    print(f"  총 010번호: {len(records)}개")
    print(f"{'='*62}")


if __name__ == "__main__":
    asyncio.run(main())
